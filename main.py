from PySide2.QtWidgets import(QApplication, QFileDialog,
 QMainWindow, QMessageBox, QTreeWidgetItem, QWidget)
from PySide2 import QtCore
from PySide2.QtGui import QIcon
from datetime import datetime
from PyQt5 import QtWidgets
from google.oauth2 import service_account
import google.oauth2.credentials as goc
import json

from google.cloud import storage
from oauth2client.service_account import ServiceAccountCredentials
import os
import sys
import openpyxl
import requests
 
from tkinter import *
from tkinter.messagebox import askyesno
from tkinter.messagebox import showerror
from tkinter.messagebox import showinfo

from time import sleep
from Scripts.ui_main import Ui_CrawTo
from Scripts.ui_login import Ui_Form
from Scripts.ui_cadastroemlotes import Ui_Cadastro
from Scripts.ui_sendavisos import Ui_Avisos
from Scripts.ui_uploadfiles import Ui_Upload
from tkinter import filedialog as fd


from hashlib import sha512


import json
from Config.sqlconfig import ConnectDB
import pathlib
from Scripts.ui_cadastroemlotes import Ui_Cadastro




class Login(QWidget, Ui_Form):
    
    def __init__(self) -> None:
        super(Login, self).__init__()
        self.setupUi(self)
        diricon = os.path.join(os.getcwd(),'Icons','faviconproexpress.png')
        self.setWindowIcon(QIcon(diricon))
        self.setWindowTitle("ControlDocs - Proexpress")

        self.EnterButton.clicked.connect(self.opensys)

    def opensys(self):


        conexao = ConnectDB().__mysql__()
        startDB = conexao.cursor()

        
        startDB.execute(f'SELECT user FROM users WHERE user = "{self.UserFrame.text()}"')
        user = startDB.fetchall()

        startDB.execute(f'SELECT cpf FROM users WHERE cpf = "{self.UserFrame.text()}"')
        cpf = startDB.fetchall()

        sleep(0.25)
        if len(user) > 0: 
            sleep(0.25)
            if self.UserFrame.text() in user[0]:
                sleep(0.25)
                try:
                    hashed_password = sha512(self.PasswordFrame.text().encode()).hexdigest()
                    checkpassword = f'SELECT password FROM users WHERE user = "{self.UserFrame.text()}"'
                    startDB.execute(checkpassword)
                    password = startDB.fetchall()
                    pass_word = password[0]
                    if hashed_password in pass_word:
                        self.w = MainWindow()
                        logintime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # registra data e hora do login

                        # atualiza a coluna login_time na tabela users com a hora de login
                        startDB.execute("UPDATE users SET login_time = %s WHERE user = %s", [logintime, self.UserFrame.text()])
                        conexao.commit()
                        self.w.show()
                        sleep(0.25)
                        self.close()

                except Exception as e:
                    showerror('Erro', 'Senha incorreta')
                    conexao.close()
                    print(e)
        elif len(cpf) > 0:
            sleep(0.25)
            if self.UserFrame.text() in cpf[0]:
                try:
                    hashed_password = sha512(self.PasswordFrame.text().encode()).hexdigest()
                    checkpassword = f'SELECT password FROM users WHERE cpf = "{self.UserFrame.text()}"'
                    startDB.execute(checkpassword)
                    password = startDB.fetchall()
                    pass_word = password[0]
                    if hashed_password in pass_word:
                        self.w = MainWindow()
                        logintime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # registra data e hora do login

                        # atualiza a coluna login_time na tabela users com a hora de login
                        startDB.execute("UPDATE users SET login_time = %s WHERE user = %s", [logintime, self.UserFrame.text()])
                        conexao.commit()
                        sleep(0.25)
                        self.w.show()
                        self.close()
                except Exception as e:
                    showerror('Erro', 'Senha incorreta')
                    conexao.close()
                    print(e)

        else:
            showerror('ERRO', 'Usuário não encontrado') 

class MainWindow(QMainWindow, Ui_CrawTo):

    def __init__(self) -> None:
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("ControlDocs - Proexpress")
        diricon = os.path.join(os.getcwd(),'Icons','faviconproexpress.png')
        self.setWindowIcon(QIcon(diricon))

        ## Janelas do Aplicativo
        self.retornar.clicked.connect(lambda: self.Pages.setCurrentWidget(self.MenuPrincipal))
        self.CadastrarUsuario.clicked.connect(lambda: self.Pages.setCurrentWidget(self.CadastroUsuario))
        self.FileUpload.clicked.connect(lambda: self.uploadfiles())
        self.CentralAvisos.clicked.connect(lambda: self.avisolot())
        self.InsertemLotes.clicked.connect(lambda: self.cadlotes())
        
        ##Cadastrar Usuário
        self.Salvar.clicked.connect(lambda: self.cadusuario())


        ## Fechar Aplicação
        self.Sair.clicked.connect(lambda: self.exitapp())
        
        


    def exitapp(self):
        answer = askyesno(title='Confirmação',message='Quer Realmente Sair?')
        if answer is True:
            conexao = ConnectDB().__mysql__()
            conexao.close()

            sys.exit()

    def cadlotes(self):

        self.w = CadastroEmLotes()
        self.w.show()

    def avisolot(self):

        self.w = CentralDeAvisos()
        self.w.show()

    def uploadfiles(self):
        self.w = Uploadfiles()
        self.w.show()

    def cadusuario(self):


        conexao = ConnectDB().__mysql__()
        startDB = conexao.cursor()

        NomeUsuario = self.lineNome.text()
        cpf = self.lineCPF.text()
        email = self.lineMails.text()
        Funcao = self.lineFuncao.text()
        matricula = self.lineMatricula.text()

        senha = self.lineSenha.text()
        confirmsenha = self.lineEdit_2.text()
        
        infotocat = [NomeUsuario, cpf, email,Funcao, matricula]
        bypassPW = [senha, confirmsenha]


        if bypassPW[0] != bypassPW[1]:
            showerror('ERRO', 'Senhas não conferem')
        else:
            hashedPW = sha512(bypassPW[0].encode()).hexdigest()
            startDB.execute(f'INSERT INTO users (user, cpf, email, password, funcao, matricula) VALUES ("{infotocat[0]}", "{infotocat[1]}", "{infotocat[2]}", "{hashedPW}", "{infotocat[3]}", "{infotocat[4]}")')
            conexao.commit()

            showinfo('SUCESSO', 'Usuário cadastrado com sucesso')

class CadastroEmLotes(QWidget, Ui_Cadastro):
    def __init__(self) -> None:
        super(CadastroEmLotes, self).__init__()
        diricon = os.path.join(os.getcwd(),'Icons','faviconproexpress.png')
        self.setWindowIcon(QIcon(diricon))
        self.setupUi(self)
        self.setWindowTitle("ControlDocs - Proexpress")

        self.BuscarPlanilha.clicked.connect(lambda: self.openfile())
        self.IniciarCadastro.clicked.connect(lambda: self.initcadastro(input_filename=self.lineEndr.text()))

    
    def openfile(self):
        input_filename = QtWidgets.QFileDialog.getOpenFileName()[0]
        self.lineEndr.setText(input_filename)
                
    def initcadastro(self, input_filename):
        conexao = ConnectDB().__mysql__()
        startDB = conexao.cursor()
        
        
        wrkbk_input = openpyxl.load_workbook(filename=input_filename)
        sheet_input = wrkbk_input.active

        for i in range(2, sheet_input.max_row+1):
            cell_obj = sheet_input.cell(row=i, column=1)
            if cell_obj.value is not None and cell_obj.value != '':
                matricula = cell_obj.value.replace(' ','')
                funcionario = sheet_input.cell(row=i, column=2).value
                CPF = sheet_input.cell(row=i, column=3).value
                funcao = sheet_input.cell(row=i, column=4).value
                try:
                    email = sheet_input.cell(row=i, column=5).value
                except:
                    email = ''
                try:
                    senha = sheet_input.cell(row=i, column=6).value
                except:
                    senha = ''

                infotocat = [matricula, funcionario, CPF, funcao, email, senha]

                if infotocat[5] is None:
                    showerror('ERRO', 'Usuário para cadastro da linha nº{} não possui senha informada'.format(i))
                    break

                elif infotocat[4] is None:
                    try:
                        hashedPW = sha512(infotocat[5].encode()).hexdigest()
                        startDB.execute(f'INSERT INTO users (user, cpf, password, funcao, matricula) VALUES ("{infotocat[1]}", "{infotocat[2]}", "{hashedPW}", "{infotocat[3]}", "{infotocat[0]}")')
                        conexao.commit()

                    except Exception as e:
                        print(e)
                else:
                    try:
                        hashedPW = sha512(senha.encode()).hexdigest()
                        startDB.execute(f'INSERT INTO users (user, cpf, email, password, funcao, matricula) VALUES ("{infotocat[1]}", "{infotocat[2]}", "{infotocat[4]}", "{hashedPW}", "{infotocat[3]}", "{infotocat[0]}")')
                        conexao.commit()

                    except Exception as e:
                        print(e)

                    
            if i == sheet_input.max_row:
                showinfo('SUCESSO', 'Usuários Cadastrados com sucesso')

class CentralDeAvisos(QWidget, Ui_Avisos):
    def __init__(self) -> None:
        super(CentralDeAvisos, self).__init__()
        diricon = os.path.join(os.getcwd(),'Icons','faviconproexpress.png')
        self.setWindowIcon(QIcon(diricon))
        self.setupUi(self)
        self.setWindowTitle("ControlDocs - Proexpress")

        self.BuscarPlanilha.clicked.connect(lambda: self.openfile())

        self.EnvioAvisos.clicked.connect(lambda: self.initcadastro(input_filename=self.lineEndr.text()))

    
    def openfile(self):
        input_filename = QtWidgets.QFileDialog.getOpenFileName()[0]
        self.lineEndr.setText(input_filename)
                
    def initcadastro(self, input_filename):
        conexao = ConnectDB().__mysql__()
        startDB = conexao.cursor()
        
        
        wrkbk_input = openpyxl.load_workbook(filename=input_filename)
        sheet_input = wrkbk_input.active

        for i in range(2, sheet_input.max_row+1):
            cell_obj = sheet_input.cell(row=i, column=1)
            if cell_obj.value is not None and cell_obj.value != '':
                aviso = cell_obj.value
                data = sheet_input.cell(row=i, column=2).value

                infotocat = [aviso, data]

                try:
                    
                    startDB.execute(f'INSERT INTO avisos (aviso, data) VALUES ("{infotocat[0]}", "{infotocat[1]}")')
                    conexao.commit()

                except Exception as e:
                    print(e)

                    
            if i == sheet_input.max_row:
                showinfo('SUCESSO', 'Avisos Enviados!')
                conexao.close()

class Uploadfiles(QWidget, Ui_Upload):

    def __init__(self) -> None:
        super(Uploadfiles, self).__init__()
        diricon = os.path.join(os.getcwd(),'Icons','faviconproexpress.png')
        self.setWindowIcon(QIcon(diricon))
        self.setupUi(self)
        self.setWindowTitle("ControlDocs - Proexpress")
        self.BuscarPlanilha.clicked.connect(lambda: self.openfile())

        self.IniciarUpload.clicked.connect(lambda: self.initupload(input_filename=self.lineEndr.text()))
    
    def openfile(self):
        input_filename = QtWidgets.QFileDialog.getOpenFileName()[0]
        self.lineEndr.setText(input_filename)
                

    def initupload(self, input_filename):
        conexao = ConnectDB().__mysql__()
        startDB = conexao.cursor()
        
        pathxlsx = pathlib.Path(input_filename).parent.resolve()

        
        wrkbk_input = openpyxl.load_workbook(filename=input_filename)
        sheet_input = wrkbk_input.active

        for i in range(2, sheet_input.max_row+1):
            cell_obj = sheet_input.cell(row=i, column=1)
            if cell_obj.value is not None and cell_obj.value != '':
                funcionario = cell_obj.value
                tipodoc = sheet_input.cell(row=i, column=2).value
                documento = sheet_input.cell(row=i, column=3).value
                mesanoreferencia = sheet_input.cell(row=i, column=4).value

                infotocat = [funcionario, tipodoc, documento, mesanoreferencia]

                if infotocat[2] is None:
                    showerror('ERRO NA LINHA Nº{}', 'É necessário informar o documento'.format(i))
                    break


                elif infotocat is not None:
                    try:
                        startDB.execute(f'INSERT INTO files (usuario, tipo_de_arquivo, data_de_inclusao, link_para_download) VALUES ("{infotocat[0]}", "{infotocat[1]}", "{infotocat[3]}", "{infotocat[2]}")')
                        conexao.commit()
                        

                        filetoupload = pathxlsx / documento
                        print(filetoupload)
                        credentials_dict = {
                        "type": "service_account",
                        "project_id": "expanded-system-379317",
                        "private_key_id": "2e2a4eb1b9eed7c425692d888cf924792c723774",
                        "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDma2WNFH2p/Q4q\nOt6P6kSXldzz+F6nbfThp3+oHs19V9KrimHqMKAjUj5Qw3LwvMVY3E7IdBmJppK9\njmHXXmroAFObBUJCP3fR+oeABIxIBla/G4bLKB7qLyCOyu8c4Lh7EK6cwA85dBC4\n2ktYfDeDCn/TA/HHfMT4sY+pkqkwddYjEf51AN07c3uxos1ycnr5aiXEo9GBUBUl\nIzHMuU/F/OE8lviEmq1lAzY3sx1b1+EhYwpLPuUt4YuNW9ZVQ3osvYVeYBxn0cEM\nJyHhZBKeR79mZ5vRc9vmwCV9HNxhhbQuv/71zbQTErJNpmlHLqU32Jw8ktaf9p7a\nmTjAIUGpAgMBAAECggEAFR5zUpS19tpfjb3TD9g+PLGK7J6PWGOy3era/7cBCRO/\nVqMqmcsmpk4fsYnUc12t7djVbLh9bbkjenW+IH4CKOz8vLjk64sDKa7/c9uW2vQf\nx8jXWC//c2Hp50SQ2XH8XqU3B91vl8m8TaEOZlgOQ3e3IwlZw/m62Pt3SijL0onR\n1mP3Me7owj6/iXh0aXCDCOysj/cwmvHbjTGUPjb6qpt70TqyN88pbUIv0oWciR4r\nSjbC44PUxzJFty82uypX8uQFPPXROug73lhlFb3fN/YwkQSuKXSMmuZq4CwkvWpt\n6pleo+sn8VKrrR0pTsMTbuehRzCXarc75H/GOKv+DQKBgQD1O1c9dbkVoZifONMy\nveyENL0F2OODMqYSk2CWRLNQNQSjnS+m2ZQwS2X0qhjAJRQB1X36RWW5P3/Ajdu1\nBj3DwfMPXJDM+Tst4pFCtRVttAK23nEwAWTiwP3EFSVbFRNJ5Vc6pv7QfPWkWS6A\nr6k+viOaGLA5Hs5yBzELosfB/QKBgQDwiYzvZFijSz6Q/B6dINeO/pCdAvT9gdGd\nF45cFLLHl/PoT+4EieI99ryWiWA6z5OUGcicN0xIbmgPyViFBTQZIijCkopl+nbg\nRxO3mISZ57+XyPpCgmipaPfSwebzkUpOthMlcecXeb7z0Usmzs1zF2TFVLie4GSH\nfLlYR2/oHQKBgQCFSS9tdCKYoy/0uknv9lIquQ9TskJV+J1S009X3RcorOhMlQ5l\nPTTR2ukbRagWJxDsvkCuWjGjseOZoICdrzq6vxuAaes5vOxxUAdrnkh7S699/QC/\n+sH+jK5geK3JB3doXoI+mq08W+6W2PHxd4ghVe67vldpdFj4mPgLPvobcQKBgQC3\nBVJRTJ8NENG4a723fSfHS/cpIPjcc4zU5PmTIsStjfcuirLmjSuyTi1MtZac6Ri3\nla7I/mJysl9Hs8JA5KaxgQZWlj3sPll5aVudVNWAbY+gpHGXbbbJQG9g090w8xoB\nLhT/Zt2ZRS+AQX3XN3+n1OL+KuqXKNIVLVexzhleGQKBgHKvsW1n1IlRtKd/mTPT\ntU5o0HDjGbfbKv2qdUNz8lmHKBBLHEU0rkdvupxQED2og9nvz2yE3Mmpsl1+xXek\n7CgSvLVKC0Rb5LJp/SeRKvx6+EqD10+HTDfnH1um7UGeSsBnkZ0mkGCaGecTEE/c\nf1yJxHpgL90it6x+he08hZh+\n-----END PRIVATE KEY-----\n",
                        "client_email": "acessorestritoarquivosmoto@expanded-system-379317.iam.gserviceaccount.com",
                        "client_id": "115060247285711623830",
                        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                        "token_uri": "https://oauth2.googleapis.com/token",
                        "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                        "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/acessorestritoarquivosmoto%40expanded-system-379317.iam.gserviceaccount.com"
                        }
                        jsonString = json.dumps(credentials_dict)
                        open 
                        try:
                            pathcredentials = pathlib.Path(__file__).parent.resolve() / 'credentials.json'
                            with open(pathcredentials, "w") as f:
                                f.write(jsonString)
                            
                        except Exception as e:
                            os.remove(pathcredentials)
                            print(e)

                        credentials = service_account.Credentials.from_service_account_file(
                        pathcredentials)
                        scoped_credentials = credentials.with_scopes(
                        ['https://www.googleapis.com/auth/cloud-platform'])
                        client = storage.Client(credentials=scoped_credentials, project='My First Project')
                        bucket = client.get_bucket('controldocs-proexpress')
                        blob = bucket.blob(str(documento))
                        blob.upload_from_filename(filetoupload)
                        f.close()
                        os.remove(pathcredentials)

                        
                    except Exception as e:
                        os.remove(pathcredentials)
                        print(e)
                else:
                    showerror('ERRO', "Falha ao enviar arquivos")
                    break
                    
            if i == sheet_input.max_row:
                showinfo('SUCESSO', 'Arquivos enviados com sucesso')
  
if __name__ == "__main__":
    
    
    app = QApplication(sys.argv)
    window = Login()
    window.show()
    app.exec_()
    